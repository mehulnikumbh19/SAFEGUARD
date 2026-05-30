from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from calculations import assign_risk_rating, compute_assessment_readiness_score, flag_overdue_remediation
from config import EXPORT_DIR
from models import Control, Evidence, Observation, PolicyMapping, Remediation, System


HEADER_FILL = PatternFill("solid", fgColor="1f2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RISK_FILLS = {
    "Critical": PatternFill("solid", fgColor="FEE2E2"),
    "High": PatternFill("solid", fgColor="FFEDD5"),
    "Medium": PatternFill("solid", fgColor="FEF3C7"),
    "Low": PatternFill("solid", fgColor="DCFCE7"),
}


def _system_rows(systems):
    return [{
        "System": s.system_name,
        "Business Owner": s.business_owner,
        "Technical Owner": s.technical_owner,
        "Environment": s.environment,
        "System Type": s.system_type,
        "Data Classification": s.data_classification,
        "Sensitive Data": "Yes" if s.sensitive_data_involved else "No",
        "Data Types": s.data_types,
        "Cloud Provider": s.cloud_provider,
        "AWS Service Type": s.aws_service_type,
        "Internet Facing": "Yes" if s.internet_facing else "No",
        "Criticality": s.criticality,
        "Assessment Status": s.assessment_status,
        "Last Reviewed": s.last_reviewed_date,
        "Next Review": s.next_review_date,
        "Notes": s.notes,
    } for s in systems]


def _control_rows(controls):
    return [{
        "System": c.system.system_name,
        "Domain": c.control_domain,
        "Control": c.control_name,
        "Objective": c.control_objective,
        "Expected Implementation": c.expected_implementation,
        "Implementation Status": c.implementation_status,
        "Owner": c.control_owner,
        "Evidence Required": "Yes" if c.evidence_required else "No",
        "Evidence Status": c.evidence_status,
        "Testing Method": c.testing_method,
        "Assessor Notes": c.assessor_notes,
        "Last Tested": c.last_tested_date,
        "Next Test": c.next_test_date,
    } for c in controls]


def _evidence_rows(evidence):
    return [{
        "System": e.system.system_name,
        "Control": e.control.control_name if e.control else "",
        "Evidence": e.evidence_name,
        "Type": e.evidence_type,
        "Source": e.evidence_source,
        "Owner": e.evidence_owner,
        "Collection Method": e.collection_method,
        "Status": e.evidence_status,
        "Date Collected": e.date_collected,
        "Expiration Date": e.expiration_date,
        "Path or Link": e.file_path_or_link,
        "Review Notes": e.review_notes,
        "Request Note": e.request_note,
    } for e in evidence]


def _observation_rows(observations):
    return [{
        "System": o.system.system_name,
        "Control": o.control.control_name if o.control else "",
        "Title": o.observation_title,
        "Risk Theme": o.risk_theme,
        "Severity": o.severity,
        "Likelihood": o.likelihood,
        "Impact": o.impact,
        "Risk Score": o.risk_score,
        "Risk Rating": assign_risk_rating(o.risk_score),
        "Business Impact": o.business_impact,
        "Compliance Impact": o.compliance_impact,
        "Recommended Action": o.recommended_action,
        "Status": o.observation_status,
        "Owner": o.owner,
        "Due Date": o.due_date,
    } for o in observations]


def _remediation_rows(remediations):
    return [{
        "Observation": r.observation.observation_title,
        "System": r.system.system_name,
        "Owner": r.remediation_owner,
        "Action Plan": r.action_plan,
        "Priority": r.priority,
        "Target Date": r.target_date,
        "Status": r.status,
        "Overdue": "Yes" if flag_overdue_remediation(r) else "No",
        "Validation Required": "Yes" if r.validation_required else "No",
        "Validation Method": r.validation_method,
        "Closure Evidence": r.closure_evidence,
        "Date Closed": r.date_closed,
    } for r in remediations]


def _mapping_rows(mappings):
    return [{
        "Control": m.control.control_name,
        "Domain": m.control.control_domain,
        "Framework": m.framework,
        "Policy Reference": m.policy_reference,
        "Requirement Summary": m.requirement_summary,
        "Mapped Objective": m.mapped_control_objective,
        "Evidence Expectation": m.evidence_expectation,
        "Notes": m.notes,
    } for m in mappings]


def _format_workbook(writer):
    workbook = writer.book
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 55)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        if worksheet.title in {"Observations", "Risk Register"}:
            for row in worksheet.iter_rows(min_row=2):
                values = {worksheet.cell(row=1, column=cell.column).value: cell for cell in row}
                risk_cell = values.get("Risk Rating")
                if risk_cell and risk_cell.value in RISK_FILLS:
                    for cell in row:
                        cell.fill = RISK_FILLS[risk_cell.value]


def export_workbook(filename="SAFEGUARD_Report.xlsx", system=None):
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORT_DIR / filename
    systems = [system] if system else System.query.order_by(System.system_name).all()
    system_ids = [s.system_id for s in systems]
    controls = Control.query.filter(Control.system_id.in_(system_ids)).all()
    evidence = Evidence.query.filter(Evidence.system_id.in_(system_ids)).all()
    observations = Observation.query.filter(Observation.system_id.in_(system_ids)).all()
    remediations = Remediation.query.filter(Remediation.system_id.in_(system_ids)).all()
    mappings = PolicyMapping.query.join(Control).filter(Control.system_id.in_(system_ids)).all()
    readiness = compute_assessment_readiness_score(systems, controls, evidence, observations, remediations)

    dashboard = [{
        "Scope": system.system_name if system else "All systems",
        "Assessment Dates": "Mar 2025 - May 2025",
        "Readiness Score": readiness,
        "Systems Reviewed": len(systems),
        "Controls": len(controls),
        "Evidence Records": len(evidence),
        "Open Observations": sum(1 for o in observations if o.observation_status not in {"Closed", "Remediated", "Risk Accepted"}),
        "Overdue Remediation": sum(1 for r in remediations if flag_overdue_remediation(r)),
    }]

    risk_register = _observation_rows(observations)
    for row in risk_register:
        remediation = next((r for r in remediations if r.observation.observation_title == row["Title"]), None)
        row["Remediation Status"] = remediation.status if remediation else ""
        row["Target Date"] = remediation.target_date if remediation else ""
        row["Days Overdue"] = "" if not remediation or not flag_overdue_remediation(remediation) else (pd.Timestamp.today().date() - remediation.target_date).days

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(dashboard).to_excel(writer, sheet_name="Cover Executive Summary", index=False)
        pd.DataFrame(_system_rows(systems)).to_excel(writer, sheet_name="Systems", index=False)
        pd.DataFrame(_control_rows(controls)).to_excel(writer, sheet_name="Controls", index=False)
        pd.DataFrame(_evidence_rows(evidence)).to_excel(writer, sheet_name="Evidence", index=False)
        pd.DataFrame(_observation_rows(observations)).to_excel(writer, sheet_name="Observations", index=False)
        pd.DataFrame(_remediation_rows(remediations)).to_excel(writer, sheet_name="Remediation", index=False)
        pd.DataFrame(_mapping_rows(mappings)).to_excel(writer, sheet_name="Policy Mapping", index=False)
        pd.DataFrame(dashboard).to_excel(writer, sheet_name="Dashboard Summary", index=False)
        pd.DataFrame(risk_register).to_excel(writer, sheet_name="Risk Register", index=False)
        _format_workbook(writer)
    return path


def export_systems_csv(systems, filename="systems_filtered.csv"):
    path = EXPORT_DIR / filename
    pd.DataFrame(_system_rows(systems)).to_csv(path, index=False)
    return path
